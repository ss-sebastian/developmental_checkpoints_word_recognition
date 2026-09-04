from __future__ import annotations

import csv
import hashlib
import html
import json
import math
import re
from collections import Counter, defaultdict
from dataclasses import dataclass
from pathlib import Path

import xlrd
from openpyxl import load_workbook


VOWEL_RE = re.compile(r"\d$")
WORD_RE = re.compile(r"^[a-z]+$")


@dataclass(frozen=True)
class Pronunciation:
    phones_stressed: tuple[str, ...]

    @property
    def phones(self) -> tuple[str, ...]:
        return tuple(re.sub(r"\d", "", phone) for phone in self.phones_stressed)

    @property
    def syllables(self) -> int:
        return sum(bool(VOWEL_RE.search(phone)) for phone in self.phones_stressed)

    @property
    def stressed_vowel_index(self) -> int:
        primary = next((i for i, phone in enumerate(self.phones_stressed) if phone.endswith("1")), None)
        if primary is not None:
            return primary
        return next(i for i, phone in enumerate(self.phones_stressed) if VOWEL_RE.search(phone))

    @property
    def onset(self) -> tuple[str, ...]:
        return self.phones[:self.stressed_vowel_index]

    @property
    def rime(self) -> tuple[str, ...]:
        return self.phones[self.stressed_vowel_index:]


@dataclass(frozen=True)
class Frequency:
    count: int
    per_million: float
    log10_count: float
    zipf: float


@dataclass(frozen=True)
class PartOfSpeech:
    dominant: str
    dominant_frequency: int
    dominant_proportion: float
    all_parts: tuple[str, ...]


@dataclass(frozen=True)
class Association:
    cue: str
    target: str
    fsg: float
    bsg: float | None
    normed_target: bool
    source_record_id: str


def sha256_file(path: str | Path) -> str:
    digest = hashlib.sha256()
    with Path(path).open("rb") as handle:
        for block in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def load_cmudict(path: str | Path) -> tuple[dict[str, Pronunciation], dict[str, int]]:
    variants: dict[str, list[Pronunciation]] = defaultdict(list)
    with Path(path).open(encoding="latin-1") as handle:
        for line in handle:
            if not line.strip() or line.startswith(";;;"):
                continue
            key, *phones = line.strip().split()
            word = re.sub(r"\(\d+\)$", "", key.lower())
            if WORD_RE.fullmatch(word):
                variants[word].append(Pronunciation(tuple(phones)))
    unambiguous = {word: values[0] for word, values in variants.items() if len(values) == 1}
    return unambiguous, {word: len(values) for word, values in variants.items()}


def load_subtlex(path: str | Path) -> tuple[dict[str, Frequency], set[str]]:
    sheet = xlrd.open_workbook(str(path), on_demand=True).sheet_by_index(0)
    header = {str(sheet.cell_value(0, column)).strip(): column for column in range(sheet.ncols)}
    required = {"Word", "FREQcount", "SUBTLWF", "Lg10WF"}
    if missing := required - header.keys():
        raise ValueError(f"SUBTLEX-US is missing columns: {sorted(missing)}")
    frequencies: dict[str, Frequency] = {}
    lowercase_entries: set[str] = set()
    for row in range(1, sheet.nrows):
        surface = str(sheet.cell_value(row, header["Word"])).strip()
        word = surface.lower()
        if not WORD_RE.fullmatch(word):
            continue
        count = int(sheet.cell_value(row, header["FREQcount"]))
        per_million = float(sheet.cell_value(row, header["SUBTLWF"]))
        log10_count = float(sheet.cell_value(row, header["Lg10WF"]))
        if word not in frequencies or surface == word:
            frequencies[word] = Frequency(count, per_million, log10_count, math.log10(per_million) + 3.0)
        if surface == word:
            lowercase_entries.add(word)
    return frequencies, lowercase_entries


def load_subtlex_pos(path: str | Path) -> dict[str, PartOfSpeech]:
    """Load the official Brysbaert-New-Keuleers SUBTLEX-US PoS extension."""
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    rows = sheet.iter_rows(values_only=True)
    header = {str(value).strip(): index for index, value in enumerate(next(rows))}
    required = {
        "Word", "Dom_PoS_SUBTLEX", "Freq_dom_PoS_SUBTLEX",
        "Percentage_dom_PoS", "All_PoS_SUBTLEX",
    }
    if missing := required - header.keys():
        raise ValueError(f"SUBTLEX-US PoS file is missing columns: {sorted(missing)}")
    result: dict[str, PartOfSpeech] = {}
    lowercase: set[str] = set()

    def numeric(value: object) -> float:
        try:
            return float(value or 0.0)
        except (TypeError, ValueError):
            return 0.0

    for row in rows:
        surface = str(row[header["Word"]] or "").strip()
        word = surface.lower()
        if not WORD_RE.fullmatch(word):
            continue
        value = PartOfSpeech(
            dominant=str(row[header["Dom_PoS_SUBTLEX"]] or "").strip(),
            dominant_frequency=int(numeric(row[header["Freq_dom_PoS_SUBTLEX"]])),
            dominant_proportion=numeric(row[header["Percentage_dom_PoS"]]),
            all_parts=tuple(str(row[header["All_PoS_SUBTLEX"]] or "").split(".")),
        )
        if word not in result or surface == word:
            result[word] = value
        if surface == word:
            lowercase.add(word)
    workbook.close()
    # Match the proper-name policy used by the base SUBTLEX loader.
    return {word: value for word, value in result.items() if word in lowercase}


def load_usf(directory: str | Path) -> tuple[list[Association], set[str]]:
    associations: list[Association] = []
    normed_cues: set[str] = set()
    for path in sorted(Path(directory).glob("Cue_Target_Pairs.*")):
        lines = path.read_text(encoding="latin-1").splitlines()
        data_lines = [html.unescape(line) for line in lines if line and not line.startswith("<")]
        for row_number, row in enumerate(csv.DictReader(data_lines, skipinitialspace=True), 2):
            cue = (row.get("CUE") or "").strip().lower()
            target = (row.get("TARGET") or "").strip().lower()
            if not WORD_RE.fullmatch(cue) or not WORD_RE.fullmatch(target):
                continue
            try:
                fsg = float((row.get("FSG") or "").strip())
            except ValueError:
                continue
            raw_bsg = (row.get("BSG") or "").strip()
            try:
                bsg = float(raw_bsg)
            except ValueError:
                bsg = None
            normed = (row.get("NORMED?") or "").strip().upper() in {"YES", "Y"}
            associations.append(Association(cue, target, fsg, bsg, normed, f"{path.name}:{row_number}"))
            normed_cues.add(cue)
    # Multiword/non-alphabetic targets are intentionally outside this project's lexical interface.
    if len(associations) < 60_000:
        raise ValueError(f"USF Appendix A parsed only {len(associations):,} usable single-word records; expected over 60,000")
    return associations, normed_cues


def normalize_sentence(parts: list[str]) -> str:
    return re.sub(r"\s+", " ", " ".join(part.strip() for part in parts if part and part.strip() != "n/a")).strip()


def verb_lemma(surface: str) -> str:
    if surface in {"is", "are", "was", "were", "be", "does", "do", "did", "not", "n/a", ""}:
        return ""
    from lemminflect import getAllLemmas
    lemmas = getAllLemmas(surface.lower()).get("VERB", ())
    return str(lemmas[0]) if lemmas else surface.lower()


@dataclass(frozen=True)
class DsExclusions:
    word_pairs: dict[str, set[tuple[str, str]]]
    task_vocabulary: dict[str, set[str]]
    sentences: dict[str, set[str]]
    critical_pairs: dict[str, set[tuple[str, str]]]
    source_rows: dict[str, list[dict[str, str]]]


def load_ds003604(directory: str | Path) -> DsExclusions:
    directory = Path(directory)
    mapping = {"Phon": "Sound", "Sem": "Meaning", "Plaus": "Plausibility", "Gram": "Grammaticality"}
    word_pairs = {task: set() for task in mapping.values()}
    task_vocabulary = {task: set() for task in mapping.values()}
    sentences = {task: set() for task in mapping.values()}
    critical_pairs = {task: set() for task in mapping.values()}
    source_rows: dict[str, list[dict[str, str]]] = {}
    for short, task in mapping.items():
        path = directory / f"task-{short}_Stimulus_Characteristics.tsv"
        with path.open(encoding="utf-8-sig", newline="") as handle:
            rows = list(csv.DictReader(handle, delimiter="\t"))
        rows = [row for row in rows if not (row.get("trial_type") or "").endswith("_C")]
        source_rows[task] = rows
        if task in {"Sound", "Meaning"}:
            for row in rows:
                pair = ((row.get("word_A") or "").strip().lower(), (row.get("word_B") or "").strip().lower())
                if all(WORD_RE.fullmatch(word) for word in pair):
                    word_pairs[task].add(pair)
                    task_vocabulary[task].update(pair)
        else:
            for row in rows:
                surfaces = [(row.get(f"verb{i}") or "").strip().lower() for i in (1, 2, 3)]
                verb = next((verb_lemma(value) for value in reversed(surfaces) if verb_lemma(value)), "")
                obj = (row.get("object") or "").strip().lower()
                sentence = normalize_sentence([
                    row.get("carrier_phrase", ""), row.get("subject", ""),
                    row.get("verb1", ""), row.get("verb2", ""), row.get("verb3", ""),
                    row.get("number", ""), row.get("object", ""),
                ]).lower()
                if sentence:
                    sentences[task].add(sentence)
                if verb and WORD_RE.fullmatch(obj):
                    critical_pairs[task].add((verb, obj))
                    task_vocabulary[task].update((verb, obj))
    return DsExclusions(word_pairs, task_vocabulary, sentences, critical_pairs, source_rows)


def count_childes_words(path: str | Path) -> tuple[Counter[str], int]:
    counts: Counter[str] = Counter()
    rows = 0
    with Path(path).open(encoding="utf-8-sig", newline="") as handle:
        for row in csv.DictReader(handle):
            rows += 1
            text = (row.get("processed_gloss") or row.get("gloss") or "").lower()
            counts.update(re.findall(r"[a-z]+(?:'[a-z]+)?", text))
    return counts, rows


def write_json(path: str | Path, value: object) -> None:
    Path(path).write_text(json.dumps(value, ensure_ascii=False, indent=2, sort_keys=True) + "\n", encoding="utf-8")
